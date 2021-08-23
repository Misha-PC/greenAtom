from config import header, column_type, file_path, sheet_name
from BaseAppException import BaseAppException
from win32com.client import Dispatch
import openpyxl.styles.numbers
import openpyxl
import os

column_type_code = {
    'finance': 44,
    'percent': 9,
    'date': 14,
    'num': 2
}


class NoSheetError(BaseAppException):
    pass


class CellFormatException(BaseAppException):
    pass


def prepear_data_to_excel(usd: list, eur: list, header: list) -> list:
    """
    :param usd: список с курсом доллора
    :param eur: список с курсом евро
    :param header: список с курсом названиями столбцов
    :return: list
    """
    output = list()
    output.append(header)

    lines_count = max(len(usd), len(eur))

    for i in range(lines_count):
        if len(usd) > i and len(eur):
            output.append(usd[i] + [''] + eur[i] + [''] + [float(eur[i][1]) / float(usd[i][1])])
        elif len(usd) > i:
            output.append(usd[i])
        else:
            output.append(eur[i])

    return [output, lines_count + 1]


def write_array_to_excel(content: list, file_path: str = file_path):
    """
    :param content: список с данными для записи в таблицу
    :param file_path: путь к файлу
    :return: None
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in content:
        ws.append(row)

    for type_ in column_type.keys():
        for col in column_type[type_]:
            for row in range(1, 30):
                style = openpyxl.styles.numbers.BUILTIN_FORMATS[column_type_code[type_]]
                ws.cell(column=col, row=row).number_format = style

    wb.save(filename=file_path)


def check_cell_type(file_pah: str):
    """
    :param file_pah:
    :return: True если тип ячеек совпадает с неоходимым в противном случае бросает исключение CellFormatException
    """
    wb = openpyxl.load_workbook(filename=file_pah)
    if not sheet_name in wb.sheetnames:
        raise NoSheetError(f"No '{sheet_name}' from '{file_pah}'")

    ws = wb[sheet_name]

    for type_ in column_type.keys():
        for col in column_type[type_]:
            for row in range(1, 30):
                style = openpyxl.styles.numbers.BUILTIN_FORMATS[column_type_code[type_]]
                if ws.cell(column=col, row=row).number_format != style:
                    raise CellFormatException(f"Cell[{col}, {row}] format is not {style}")
    return True


def auto_fit(file_path: str = file_path):
    """
    Устанавливает автоширину для столбцов A:G
    :param file_path:
    :return:
    """
    full_path = os.path.abspath(file_path)
    xl_app = Dispatch('Excel.Application')
    xl_app.Visible = True
    wb = xl_app.Workbooks.Open(full_path)
    wb.ActiveSheet.Columns("A:G").AutoFit()
    xl_app.ActiveWorkbook.Close(SaveChanges=1)
    xl_app.Quit()
    del xl_app
